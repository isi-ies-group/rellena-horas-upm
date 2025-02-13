# -*- coding: utf-8 -*-
"""
Created on Tue Apr 13 14:09:09 2021

@author: Ruben
"""
import openpyxl
import openpyxl.utils
from datetime import datetime, timedelta
import numpy as np

# %% ENTRADAS
nombre_archivo = '2025-02-03_1300_carga_horas_2023_c-FIVe.xlsx'

# array de tamaño NUM_TAREAS indicando el número de horas de cada tarea a asignar
HORAS_TASK = np.array([0, 0, 0, 0, 0, 0, 50, 0, 50, 0])
# OJO han de ser floats (terminados con .0), se fuerza con astype()
HORAS_TASK = HORAS_TASK.astype('float')

# meses en los cuales se imputan horas a las tareas
MESES_IMPUTAR = [True, True, True, True, True, True, True, False, True, True, True, True]

# lista días laborables a excluir
EXCLUYE_INI = '01/01/2020'
EXCLUYE_FIN = '01/01/2020'

exc_ini = datetime.strptime(EXCLUYE_INI, "%d/%m/%Y")
exc_fin = datetime.strptime(EXCLUYE_FIN, "%d/%m/%Y")
DIAS_EXCLUIR = [(exc_ini + timedelta(days=x)).strftime("%d/%m/%Y") for x in range(0, (exc_fin-exc_ini).days)]
# DIAS_EXCLUIR = ['02/03/2020', '03/03/2020'] # excluir días sueltos

INICIALIZAR_DIAS = True # pone a cero las horas de días disponibles antes de repartir las horas ahora

# %% CONSTANTES
# fijar la semilla del random permite reproducir siempre el mismo patrón aleatorio
np.random.seed(123)

COLOR_FESTIVO = 52 # MOSTAZA
COLOR_BLOQUEADO = 10 # ROJO
COLOR_DISPONIBLE = 9 # BLANCO

HORAS_DIA_MAX = 7.5
MINIMA_CARGA_HORARIA = 0.5

COL_INI_TAREA = 'B'
COL_FECHA = 'A'
FILA_CABECERAS_TAREAS = '3'
FILA_INI_TAREAS = '4'

OFFSET_COLUMNA_TAREA = openpyxl.utils.column_index_from_string(COL_INI_TAREA)

wb = openpyxl.load_workbook(nombre_archivo)

hoja = wb[wb.sheetnames[0]]

NUM_FILAS = len(hoja[COL_INI_TAREA]) - 1 # la última fila es de subtotales. Si se incluye se machaca al rellenar con 0s

for celda in hoja[FILA_CABECERAS_TAREAS]:
    if celda.value == 'Otras actividades':
        COL_OTRAS_IDX = celda.col_idx

COL_OTRAS_TAREAS = openpyxl.utils.get_column_letter(COL_OTRAS_IDX+2)

COL_FIN_TAREA = openpyxl.utils.get_column_letter(COL_OTRAS_IDX-1)

# %% Lee los días laborables
# lista_dias = []
lista_filas = []
lista_horas_otros_proyectos = []

for dia in hoja[f'{COL_INI_TAREA}{FILA_INI_TAREAS}:B{NUM_FILAS}']:
    for celda in dia:
        celda_fecha = hoja[COL_FECHA+str(celda.row)]
        if celda.fill.fgColor.indexed == COLOR_DISPONIBLE and celda_fecha.value is not None:
            if MESES_IMPUTAR[datetime.strptime(celda_fecha.value, '%d/%m/%Y').date().month - 1]:
                if celda_fecha.value not in DIAS_EXCLUIR:
                    lista_filas.append(celda_fecha.row)
                    # lista_dias.append(celda_fecha.value)
                    lista_horas_otros_proyectos.append(
                        hoja[f'{COL_OTRAS_TAREAS}{celda_fecha.row}'].value)

if len(lista_horas_otros_proyectos)*HORAS_DIA_MAX - sum(lista_horas_otros_proyectos) < sum(HORAS_TASK):
    raise SystemError('Más horas en tareas que horas disponibles!')

# %% Inicializa a 0 todas las tareas de todos los días laborables
if INICIALIZAR_DIAS:
    for dia in hoja[f'{COL_INI_TAREA}{FILA_INI_TAREAS}:B{NUM_FILAS}']:
        for celda in dia:
            if celda.fill.fgColor.indexed == COLOR_DISPONIBLE:
                for tarea in range(OFFSET_COLUMNA_TAREA, COL_OTRAS_IDX):
                    hoja[f'{openpyxl.utils.get_column_letter(tarea)}{dia[0].row}'].value = 0


# %% Rellena la hoja con horas en las tareas
horas_task_ahora = HORAS_TASK.copy()

lista_filas_random = np.array(lista_filas)
# OJO hace el shuffle in_place!
np.random.shuffle(lista_filas_random)

while horas_task_ahora.sum() > 0:
    for num_dia in lista_filas_random:
        for tarea in range(OFFSET_COLUMNA_TAREA, COL_OTRAS_IDX):

            horas_proyectos_ahora = hoja[f'{COL_INI_TAREA}{num_dia}:{COL_FIN_TAREA}{num_dia}']
            total_proyectos_ahora = sum([c.value for c in list(horas_proyectos_ahora[0])])

            total_otros_proyectos = hoja[f'{COL_OTRAS_TAREAS}{num_dia}'].value

            horas_libres_dia = HORAS_DIA_MAX - total_proyectos_ahora - total_otros_proyectos

            if horas_libres_dia > 0 and horas_task_ahora[tarea-OFFSET_COLUMNA_TAREA] > 0:
                hoja[f'{openpyxl.utils.get_column_letter(tarea)}{num_dia}'].value += MINIMA_CARGA_HORARIA
                horas_task_ahora[tarea-OFFSET_COLUMNA_TAREA] -= MINIMA_CARGA_HORARIA

# %% Guarda en copia
nombre_partido = nombre_archivo.split('.')
nombre_archivo_relleno = nombre_partido[0] + '_relleno' + '.' + nombre_partido[1]

wb.save(nombre_archivo_relleno)

print('Generado fichero relleno:', nombre_archivo_relleno)
